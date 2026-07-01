import streamlit as st
import geopandas as gpd
import pandas as pd
import fiona
import folium
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import zipfile, os, io, tempfile
from datetime import datetime

# ── Configuração da página ─────────────────────────────────────
st.set_page_config(
    page_title="Comparador de Redes — Águas do Rio",
    page_icon="🗺️",
    layout="wide",
)

# ── Estilo visual ──────────────────────────────────────────────
st.markdown("""
<style>
    .main { background-color: #F8FAFC; }
    .stApp { font-family: Arial, sans-serif; }
    .titulo { color: #1F4E79; font-size: 2rem; font-weight: bold; }
    .subtitulo { color: #2E75B6; font-size: 1rem; margin-bottom: 1.5rem; }
    .card {
        background: white;
        border-radius: 10px;
        padding: 1.2rem 1.5rem;
        border: 1px solid #E0E8F0;
        margin-bottom: 1rem;
    }
    .stat-num { font-size: 2.2rem; font-weight: bold; }
    .stat-label { font-size: 0.85rem; color: #666; }
    .rem  { color: #E74C3C; }
    .add  { color: #27AE60; }
    .alt  { color: #F39C12; }
    .ok   { color: #1F4E79; }
</style>
""", unsafe_allow_html=True)

# ── Cabeçalho ─────────────────────────────────────────────────
st.markdown('<div class="titulo">🗺️ Comparador de Redes Geoespaciais</div>', unsafe_allow_html=True)
st.markdown('<div class="subtitulo">Águas do Rio | Planejamento e Controle — Detecta diferenças entre duas versões de uma camada vetorial</div>', unsafe_allow_html=True)
st.divider()


# ── Funções auxiliares ─────────────────────────────────────────

def ler_arquivo(arquivo_bytes, nome_arquivo):
    """Lê .gpkg, .shp ou .zip contendo gpkg/shp. Retorna (gdf, camadas_disponíveis, caminho_tmp)."""
    ext = os.path.splitext(nome_arquivo)[1].lower()
    tmp_dir = tempfile.mkdtemp()

    if ext == '.zip':
        with zipfile.ZipFile(io.BytesIO(arquivo_bytes)) as z:
            z.extractall(tmp_dir)
        # Procurar gpkg ou shp dentro
        caminho = None
        for root, _, fs in os.walk(tmp_dir):
            for f in fs:
                if f.endswith('.gpkg'):
                    caminho = os.path.join(root, f)
                    break
            if caminho:
                break
        if not caminho:
            for root, _, fs in os.walk(tmp_dir):
                for f in fs:
                    if f.endswith('.shp'):
                        caminho = os.path.join(root, f)
                        break
                if caminho:
                    break
        if not caminho:
            raise ValueError("Nenhum .gpkg ou .shp encontrado dentro do ZIP.")
    elif ext == '.gpkg':
        caminho = os.path.join(tmp_dir, nome_arquivo)
        with open(caminho, 'wb') as f:
            f.write(arquivo_bytes)
    elif ext == '.shp':
        caminho = os.path.join(tmp_dir, nome_arquivo)
        with open(caminho, 'wb') as f:
            f.write(arquivo_bytes)
    else:
        raise ValueError(f"Formato não suportado: {ext}")

    ext_real = os.path.splitext(caminho)[1].lower()
    if ext_real == '.gpkg':
        camadas = fiona.listlayers(caminho)
    else:
        camadas = [os.path.basename(caminho)]

    return caminho, camadas, ext_real


def carregar_camada(caminho, ext, camada):
    if ext == '.gpkg':
        return gpd.read_file(caminho, layer=camada)
    else:
        return gpd.read_file(caminho)


def comparar(gdf_a, gdf_n, campo_id):
    if gdf_a.crs != gdf_n.crs:
        gdf_n = gdf_n.to_crs(gdf_a.crs)

    if campo_id and campo_id in gdf_a.columns and campo_id in gdf_n.columns:
        col_id = campo_id
        idx_a = set(gdf_a[col_id].astype(str))
        idx_n = set(gdf_n[col_id].astype(str))
    else:
        gdf_a = gdf_a.copy()
        gdf_n = gdf_n.copy()
        gdf_a['_gh'] = gdf_a.geometry.apply(lambda g: str(hash(g.wkt)) if g else None)
        gdf_n['_gh'] = gdf_n.geometry.apply(lambda g: str(hash(g.wkt)) if g else None)
        col_id = '_gh'
        idx_a = set(gdf_a[col_id])
        idx_n = set(gdf_n[col_id])

    ids_rem = idx_a - idx_n
    ids_add = idx_n - idx_a
    ids_com = idx_a & idx_n

    removidos   = gdf_a[gdf_a[col_id].isin(ids_rem)].copy()
    adicionados = gdf_n[gdf_n[col_id].isin(ids_add)].copy()

    alterados_lista = []
    if ids_com and col_id != '_gh':
        df_a = gdf_a[gdf_a[col_id].isin(ids_com)].set_index(col_id)
        df_n = gdf_n[gdf_n[col_id].isin(ids_com)].set_index(col_id)
        cols = [c for c in df_a.columns if c in df_n.columns and c != 'geometry']
        for idx in ids_com:
            try:
                ra, rn = df_a.loc[idx], df_n.loc[idx]
                diffs = [f"{c}: [{ra[c]}] → [{rn[c]}]" for c in cols if str(ra.get(c,'')) != str(rn.get(c,''))]
                if diffs:
                    alterados_lista.append({
                        col_id: idx,
                        '_alteracoes': ' | '.join(diffs),
                        'geometry': df_n.loc[idx, 'geometry'] if 'geometry' in df_n.columns else None
                    })
            except Exception:
                continue

    alterados = gpd.GeoDataFrame(alterados_lista, crs=gdf_a.crs) if alterados_lista else gpd.GeoDataFrame()

    return {
        'removidos': removidos,
        'adicionados': adicionados,
        'alterados': alterados,
        'total_a': len(gdf_a),
        'total_n': len(gdf_n),
    }


def validar_geometrias(gdf, label):
    problemas = []
    for i, row in gdf.iterrows():
        g = row.geometry
        if g is None or g.is_empty:
            problemas.append({'Índice': i, 'Problema': 'Geometria nula/vazia', 'Arquivo': label})
        elif not g.is_valid:
            problemas.append({'Índice': i, 'Problema': 'Geometria inválida', 'Arquivo': label})
    return pd.DataFrame(problemas)


def gerar_mapa(gdf_novo, resultado):
    gdf_ref = gdf_novo.to_crs(epsg=4326) if gdf_novo.crs and gdf_novo.crs.to_epsg() != 4326 else gdf_novo
    centro = [gdf_ref.geometry.centroid.y.mean(), gdf_ref.geometry.centroid.x.mean()]
    m = folium.Map(location=centro, zoom_start=13, tiles='CartoDB positron')

    def to_wgs(gdf):
        if gdf is None or len(gdf) == 0:
            return None
        return gdf.to_crs(epsg=4326) if gdf.crs and gdf.crs.to_epsg() != 4326 else gdf

    # Base cinza
    folium.GeoJson(to_wgs(gdf_novo).__geo_interface__,
        name='Rede atual',
        style_function=lambda f: {'color': '#AAAAAA', 'weight': 1, 'opacity': 0.4}
    ).add_to(m)

    # Removidos
    r = to_wgs(resultado['removidos'])
    if r is not None and len(r) > 0:
        folium.GeoJson(r.__geo_interface__, name=f"🔴 Removidos ({len(r)})",
            style_function=lambda f: {'color': '#E74C3C', 'weight': 3}
        ).add_to(m)

    # Adicionados
    a = to_wgs(resultado['adicionados'])
    if a is not None and len(a) > 0:
        folium.GeoJson(a.__geo_interface__, name=f"🟢 Adicionados ({len(a)})",
            style_function=lambda f: {'color': '#27AE60', 'weight': 3}
        ).add_to(m)

    # Alterados
    al = to_wgs(resultado['alterados'])
    if al is not None and len(al) > 0:
        folium.GeoJson(al.__geo_interface__, name=f"🟡 Alterados ({len(al)})",
            style_function=lambda f: {'color': '#F39C12', 'weight': 3}
        ).add_to(m)

    n_rem = len(resultado['removidos'])
    n_add = len(resultado['adicionados'])
    n_alt = len(resultado['alterados'])
    agora = datetime.now().strftime('%d/%m/%Y %H:%M')

    legenda = f"""
    <div style="position:fixed;bottom:30px;left:30px;z-index:1000;
         background:white;padding:14px;border-radius:8px;
         border:2px solid #ccc;font-family:Arial;font-size:13px;">
      <b>🗺️ Comparação de Redes</b><br>
      <span style="font-size:11px;color:#888">{agora}</span><br><br>
      <span style="color:#AAAAAA">━━</span> Rede atual<br>
      <span style="color:#E74C3C">━━</span> Removidos: <b>{n_rem}</b><br>
      <span style="color:#27AE60">━━</span> Adicionados: <b>{n_add}</b><br>
      <span style="color:#F39C12">━━</span> Alterados: <b>{n_alt}</b>
    </div>"""
    m.get_root().html.add_child(folium.Element(legenda))
    folium.LayerControl().add_to(m)
    return m


def gerar_excel(resultado, problemas_geom, nome_a, nome_n, tipo_rede, municipio):
    wb = openpyxl.Workbook()
    AZUL = '1F4E79'
    CORES = {'rem': 'FADADD', 'add': 'D5F5E3', 'alt': 'FEF9E7', 'prob': 'FDEBD0'}

    def h(cell):
        cell.font = Font(name='Arial', bold=True, color='FFFFFF')
        cell.fill = PatternFill('solid', start_color=AZUL)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        b = Side(style='thin', color='BFBFBF')
        cell.border = Border(left=b, right=b, top=b, bottom=b)

    def aba(titulo, gdf, cor, extras=None):
        ws = wb.create_sheet(titulo)
        if gdf is None or len(gdf) == 0:
            ws['A1'] = 'Nenhuma feição nesta categoria.'
            ws['A1'].font = Font(italic=True, color='888888')
            return
        cols = [c for c in gdf.columns if c not in ('geometry', '_gh', '_geom_hash')]
        if extras:
            cols = [c for c in cols if c not in extras] + extras
        for j, c in enumerate(cols, 1):
            cell = ws.cell(1, j, c.upper().replace('_', ' '))
            h(cell)
            ws.column_dimensions[get_column_letter(j)].width = max(15, len(c) + 4)
        fill = PatternFill('solid', start_color=cor)
        b = Side(style='thin', color='BFBFBF')
        borda = Border(left=b, right=b, top=b, bottom=b)
        for i, (_, row) in enumerate(gdf.iterrows(), 2):
            for j, c in enumerate(cols, 1):
                val = row.get(c, '')
                if hasattr(val, 'item'): val = val.item()
                cell = ws.cell(i, j, str(val) if val is not None else '')
                cell.font = Font(name='Arial', size=10)
                cell.fill = fill
                cell.border = borda
                cell.alignment = Alignment(vertical='center')
        ws.freeze_panes = 'A2'
        ws.auto_filter.ref = ws.dimensions

    agora = datetime.now().strftime('%d/%m/%Y %H:%M')
    n_rem = len(resultado['removidos'])
    n_add = len(resultado['adicionados'])
    n_alt = len(resultado['alterados'])

    ws = wb.active
    ws.title = 'Resumo'
    linhas = [
        ['COMPARAÇÃO DE REDES — ÁGUAS DO RIO'], [],
        ['Rede:', tipo_rede], ['Município:', municipio],
        ['Data da análise:', agora],
        ['Arquivo antigo:', nome_a], ['Arquivo novo:', nome_n], [],
        ['RESULTADO', 'QUANTIDADE'],
        ['Feições no arquivo antigo', resultado['total_a']],
        ['Feições no arquivo novo',   resultado['total_n']],
        ['🔴 Removidas', n_rem], ['🟢 Adicionadas', n_add],
        ['🟡 Atributos alterados', n_alt],
        ['⚠️ Problemas de geometria', len(problemas_geom)], [],
        ['Total de mudanças', n_rem + n_add + n_alt],
    ]
    for i, linha in enumerate(linhas, 1):
        for j, val in enumerate(linha, 1):
            cell = ws.cell(i, j, val)
            cell.font = Font(name='Arial', size=11)
            if i == 1: cell.font = Font(name='Arial', bold=True, size=14, color=AZUL)
            elif i == 9: h(cell)
            elif i == 17 and j == 2: cell.font = Font(name='Arial', bold=True, size=13)
    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 40

    aba('🔴 Removidos',   resultado['removidos'],   CORES['rem'])
    aba('🟢 Adicionados', resultado['adicionados'], CORES['add'])
    aba('🟡 Alterados',   resultado['alterados'],   CORES['alt'], extras=['_alteracoes'])
    aba('⚠️ Geometrias',  problemas_geom if len(problemas_geom) > 0 else None, CORES['prob'])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ══════════════════════════════════════════════════════════════
# INTERFACE PRINCIPAL
# ══════════════════════════════════════════════════════════════

# ── Configurações ──
with st.expander("⚙️ Configurações", expanded=True):
    col1, col2, col3 = st.columns(3)
    with col1:
        tipo_rede = st.selectbox("Tipo de rede", ["Água", "Esgoto", "Água e Esgoto", "Outro"])
    with col2:
        municipio = st.text_input("Município / Bloco", value="Rio de Janeiro")
    with col3:
        campo_id = st.text_input("Campo ID único (opcional)", placeholder="Ex: OBJECTID, cod_trecho",
                                  help="Campo que identifica unicamente cada feição. Se vazio, usa hash da geometria.")

# ── Upload ──
st.subheader("📂 Upload dos arquivos")
st.caption("Formatos aceitos: **.gpkg** · **.zip** (shapefile ou gpkg dentro)")

col_a, col_n = st.columns(2)
with col_a:
    st.markdown("**Versão ANTIGA**")
    arq_antigo = st.file_uploader("Arquivo antigo", type=['gpkg', 'zip', 'shp'],
                                   label_visibility='collapsed', key='antigo')
with col_n:
    st.markdown("**Versão NOVA**")
    arq_novo = st.file_uploader("Arquivo novo", type=['gpkg', 'zip', 'shp'],
                                 label_visibility='collapsed', key='novo')

# ── Seleção de camadas (se gpkg com múltiplas camadas) ──
camada_a = camada_n = None
caminho_a = caminho_n = None
ext_a = ext_n = None

if arq_antigo:
    try:
        caminho_a, camadas_a, ext_a = ler_arquivo(arq_antigo.read(), arq_antigo.name)
        if len(camadas_a) > 1:
            camada_a = st.selectbox(f"Camada do arquivo ANTIGO ({arq_antigo.name})", camadas_a)
        else:
            camada_a = camadas_a[0]
        st.success(f"✅ Antigo: **{arq_antigo.name}** — camada: `{camada_a}`")
    except Exception as e:
        st.error(f"Erro ao ler arquivo antigo: {e}")

if arq_novo:
    try:
        caminho_n, camadas_n, ext_n = ler_arquivo(arq_novo.read(), arq_novo.name)
        if len(camadas_n) > 1:
            camada_n = st.selectbox(f"Camada do arquivo NOVO ({arq_novo.name})", camadas_n)
        else:
            camada_n = camadas_n[0]
        st.success(f"✅ Novo: **{arq_novo.name}** — camada: `{camada_n}`")
    except Exception as e:
        st.error(f"Erro ao ler arquivo novo: {e}")

# ── Botão comparar ──
st.divider()
btn = st.button("🔍 Comparar arquivos", type="primary",
                 disabled=not (arq_antigo and arq_novo and caminho_a and caminho_n))

if btn:
    with st.spinner("Carregando e comparando..."):
        try:
            gdf_a = carregar_camada(caminho_a, ext_a, camada_a)
            gdf_n = carregar_camada(caminho_n, ext_n, camada_n)

            resultado = comparar(gdf_a, gdf_n, campo_id.strip() or None)
            val_a = validar_geometrias(gdf_a, arq_antigo.name)
            val_n = validar_geometrias(gdf_n, arq_novo.name)
            problemas_geom = pd.concat([val_a, val_n], ignore_index=True)

            n_rem = len(resultado['removidos'])
            n_add = len(resultado['adicionados'])
            n_alt = len(resultado['alterados'])

            # ── Cards de resultado ──
            st.subheader("📊 Resultado")
            c1, c2, c3, c4, c5 = st.columns(5)
            c1.metric("Feições antigas", resultado['total_a'])
            c2.metric("Feições novas",   resultado['total_n'])
            c3.metric("🔴 Removidas",    n_rem)
            c4.metric("🟢 Adicionadas",  n_add)
            c5.metric("🟡 Alteradas",    n_alt)

            if len(problemas_geom) > 0:
                st.warning(f"⚠️ {len(problemas_geom)} problema(s) de geometria encontrado(s)")
            else:
                st.success("✅ Nenhum problema de geometria encontrado.")

            # ── Tabelas detalhadas ──
            if n_rem > 0 or n_add > 0 or n_alt > 0:
                st.subheader("📋 Detalhes das mudanças")
                tab1, tab2, tab3 = st.tabs([f"🔴 Removidos ({n_rem})",
                                             f"🟢 Adicionados ({n_add})",
                                             f"🟡 Alterados ({n_alt})"])
                with tab1:
                    if n_rem > 0:
                        cols = [c for c in resultado['removidos'].columns if c not in ('geometry','_gh')]
                        st.dataframe(resultado['removidos'][cols], use_container_width=True)
                    else:
                        st.info("Nenhuma feição removida.")
                with tab2:
                    if n_add > 0:
                        cols = [c for c in resultado['adicionados'].columns if c not in ('geometry','_gh')]
                        st.dataframe(resultado['adicionados'][cols], use_container_width=True)
                    else:
                        st.info("Nenhuma feição adicionada.")
                with tab3:
                    if n_alt > 0:
                        cols = [c for c in resultado['alterados'].columns if c not in ('geometry','_gh')]
                        st.dataframe(resultado['alterados'][cols], use_container_width=True)
                    else:
                        st.info("Nenhum atributo alterado.")

            # ── Mapa ──
            st.subheader("🗺️ Mapa de diferenças")
            mapa = gerar_mapa(gdf_n, resultado)
            buf_mapa = io.BytesIO()
            mapa.save(buf_mapa, close_file=False)
            st.components.v1.html(mapa._repr_html_(), height=500)

            # ── Downloads ──
            st.subheader("📥 Downloads")
            col_d1, col_d2 = st.columns(2)

            xlsx_buf = gerar_excel(resultado, problemas_geom,
                                    arq_antigo.name, arq_novo.name,
                                    tipo_rede, municipio)
            data_str = datetime.now().strftime('%Y%m%d_%H%M')

            with col_d1:
                st.download_button(
                    "📊 Baixar relatório Excel",
                    data=xlsx_buf,
                    file_name=f"relatorio_redes_{data_str}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True,
                )
            with col_d2:
                st.download_button(
                    "🗺️ Baixar mapa HTML",
                    data=buf_mapa.getvalue(),
                    file_name=f"mapa_diferencas_{data_str}.html",
                    mime="text/html",
                    use_container_width=True,
                )

        except Exception as e:
            st.error(f"Erro durante a comparação: {e}")
            st.exception(e)

# ── Rodapé ──
st.divider()
st.caption("Águas do Rio | Planejamento e Controle · Comparador de Redes Geoespaciais")
